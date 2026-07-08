#!/usr/bin/env python3
"""
Sync MARDET data from Google Sheets (or local CSV exports) into website files.

Data sources:
  - POA&M workbook (internal ops — not synced directly; export curated subsets)
  - Marines tab → data/marines.csv → js/intranet/marines-on-the-yard-data.js
  - Company mentors (from Marines sheet or separate tab) → js/intranet/company-mentors-data.js

Usage:
  python3 scripts/sync-from-sheets.py
  python3 scripts/sync-from-sheets.py --dry-run
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
SCRIPTS_DIR = ROOT / "scripts"
STATE_FILE = DATA_DIR / ".last-sync.json"
REPORT_FILE = DATA_DIR / "sync-report.txt"
INTRANET_JS_DIR = ROOT / "js" / "intranet"
MENTOR_PHOTO_PREFIX = "../../assets/images/intranet/mentors"

COMMUNITY_SLUGS = {
    "mardet": "mardet",
    "marine detachment": "mardet",
    "detachment": "mardet",
    "mentor": "mentor",
    "company mentor": "mentor",
    "marine mentor": "mentor",
    "company officer": "company-officer",
    "co": "company-officer",
    "instructor": "instructor",
    "academic": "instructor",
    "faculty": "instructor",
    "coach": "coach",
    "athletics": "coach",
    "extracurricular": "coach",
    "staff": "support",
    "support": "support",
    "commandant": "support",
}

COMMUNITY_LABELS = {
    "mardet": "Marine Detachment (MARDET)",
    "mentor": "Company Marine Mentor",
    "company-officer": "Company Officer",
    "instructor": "Academic Instructor",
    "coach": "Athletics & Extracurricular",
    "support": "Commandant's Staff & Support",
    "other": "Other Marines on the Yard",
}

WORKBOOK_FILENAME = "MARDET_Ops_Master_Working_2026.xlsx"
MARINES_TAB = "MARDET Marines"

COMPANY_OFFICER_RE = re.compile(
    r"company\s+(?:officer|co|sel|senior\s+enlisted|sel\.?)|battalion\s+officer|brigade\s+sgtmaj",
    re.I,
)
COMPANY_FROM_BILLET_RE = re.compile(r"(\d+)\s*(?:st|nd|rd|th)?\s*(?:co|company)\b", re.I)
COMPANY_OFFICER_BILLET_RE = re.compile(r"company\s+(?:officer|co)\b", re.I)
INSTRUCTOR_RE = re.compile(r"\binstr(?:uctor)?\b", re.I)
DEPARTED_RE = re.compile(r"\bdeparted\b|\bpcs/pca/seperating\b", re.I)


def load_config() -> dict[str, str]:
    config: dict[str, str] = {}
    env_path = SCRIPTS_DIR / "config.env"
    if not env_path.exists():
        return config
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        config[key.strip()] = value.strip()
    return config


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def normalize_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]", "", key.lower())


def pick_field(row: dict[str, str], *candidates: str) -> str:
    norm_row = {normalize_key(k): ("" if v is None else str(v).strip()) for k, v in row.items()}
    for cand in candidates:
        val = norm_row.get(normalize_key(cand), "")
        if val:
            return val
    return ""


def pick_name(row: dict[str, str]) -> str:
    full = pick_field(
        row,
        "name",
        "full_name",
        "Full Name",
        "Marine",
        "LName/FName/MI",
        "LName FName MI",
        "Last, First",
    )
    if full:
        return full
    first = pick_field(row, "first_name", "First Name", "First")
    last = pick_field(row, "last_name", "Last Name", "Last")
    return f"{first} {last}".strip()


def parse_company(value: str) -> int:
    raw = (value or "").strip()
    if not raw or DEPARTED_RE.search(raw) or raw.upper().startswith("N/A"):
        return 0
    match = re.match(r"(\d+)", raw.replace(".0", ""))
    return int(match.group(1)) if match else 0


def company_from_billet(billet: str) -> int:
    match = COMPANY_FROM_BILLET_RE.search(billet or "")
    return int(match.group(1)) if match else 0


def ordinal_company(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def is_company_officer_billet(billet: str) -> bool:
    upper = (billet or "").upper()
    if any(token in upper for token in ("COMPANY SEL", "SENIOR ENLISTED", "SGTMAJ", "BATTALION OFFICER")):
        return False
    return bool(COMPANY_OFFICER_BILLET_RE.search(upper))


def is_instructor_billet(billet: str) -> bool:
    upper = (billet or "").upper()
    if "PHYSICAL EDUCATION INST" in upper or "PLAYER DEV" in upper:
        return False
    if re.search(r"\bINSTR(?:UCTOR)?\b", upper):
        return True
    return bool(re.search(r"\b(?:LEADERSHIP|AERONAUTICAL ENGR|ENGR) INST\b", upper))


def title_case_subject(subject: str) -> str:
    text = subject.strip().lower()
    text = re.sub(r"\bwpns\b", "weapons", text)
    text = re.sub(r"\bengr\b", "engineering", text)
    small_words = {"and", "or", "of", "the", "a", "an"}
    words = re.split(r"(\s+)", text)
    result: list[str] = []
    word_index = 0
    for token in words:
        if not token.strip():
            result.append(token)
            continue
        if word_index > 0 and token in small_words:
            result.append(token)
        else:
            result.append(token.capitalize())
        word_index += 1
    return "".join(result).strip(" ,")


UPPER_ACRONYMS = {
    "magtf",
    "mciws",
    "jmot",
    "mwtc",
    "mcmap",
    "mat",
    "pro",
    "mardet",
    "oic",
    "o/e",
    "fb",
    "d3",
    "i",
    "ii",
    "iii",
    "iv",
}

SMALL_WORDS = {"and", "or", "of", "the", "a", "an", "for", "to"}


def format_token(token: str, word_index: int) -> str:
    if not token:
        return token
    if token in ",/()":
        return token
    if len(token) > 2 and token.startswith("(") and token.endswith(")"):
        return f"({format_display_text(token[1:-1])})"
    if "/" in token and token.lower() not in UPPER_ACRONYMS:
        return "/".join(format_token(part, word_index) for part in token.split("/"))

    lower = token.lower()
    trailing = ""
    while lower and not lower[-1].isalnum():
        trailing = lower[-1] + trailing
        lower = lower[:-1]
    leading = ""
    while lower and not lower[0].isalnum():
        leading += lower[0]
        lower = lower[1:]

    if lower in UPPER_ACRONYMS:
        formatted = lower.upper()
    elif lower in {"jr", "sr"}:
        formatted = lower.capitalize()
    elif lower in SMALL_WORDS and word_index > 0 and len(lower) > 1:
        formatted = lower
    else:
        formatted = lower.capitalize()

    return f"{leading}{formatted}{trailing}"


def format_display_text(text: str) -> str:
    text = text.strip()
    if not text:
        return ""

    def fix_segment(segment: str, word_index: int) -> str:
        if segment.startswith("(") and segment.endswith(")"):
            return f"({format_display_text(segment[1:-1])})"
        comma = "," if segment.endswith(",") else ""
        body = segment[:-1] if comma else segment
        if "/" in body:
            parts = body.split("/")
            formatted = [fix_segment(part, word_index + index) for index, part in enumerate(parts)]
            return "/".join(formatted) + comma
        return format_token(body, word_index) + comma

    words = text.split()
    formatted_words: list[str] = []
    word_index = 0
    for word in words:
        formatted_words.append(fix_segment(word, word_index))
        if re.search(r"[A-Za-z]", word):
            word_index += 1
    return " ".join(formatted_words)


def format_rank(rank: str) -> str:
    mapping = {
        "CAPT": "Capt",
        "MAJ": "Maj",
        "LTCOL": "LtCol",
        "COL": "Col",
        "GYSGT": "GySgt",
        "MSGT": "MSgt",
        "SSGT": "SSgt",
        "1STLT": "1stLt",
        "2NDLT": "2ndLt",
    }
    key = rank.strip().upper()
    return mapping.get(key, format_display_text(rank))


def extract_instructor_type(billet: str) -> str:
    cleaned = re.sub(r"\s*\(.*?\)\s*", " ", billet).strip()
    lang_match = re.match(r"^LANGUAGE\s+INSTR(?:UCTOR)?\s*,\s*(.+)$", cleaned, re.I)
    if lang_match:
        return f"{title_case_subject(lang_match.group(1))} Language"
    subject = re.sub(r"\s+INSTR(?:UCTOR)?\s*,?\s*.*$", "", cleaned, flags=re.I)
    subject = re.sub(r"\s+INST\s*$", "", subject, flags=re.I)
    return title_case_subject(subject)


def format_primary_duty(billet: str, fallback_company: int = 0) -> str:
    billet = (billet or "").strip()
    if not billet:
        return ""
    if is_company_officer_billet(billet):
        company_num = company_from_billet(billet) or fallback_company
        if company_num:
            return f"Company Officer, {ordinal_company(company_num)} Co"
        return "Company Officer"
    if is_instructor_billet(billet):
        instructor_type = extract_instructor_type(billet)
        if instructor_type:
            return f"Instructor, {instructor_type}"
    return format_display_text(billet)


def pick_collateral_duty(row: dict[str, str]) -> str:
    return format_display_text(
        pick_field(
            row,
            "Collateral Duty '26",
            "Collateral Duty",
            "collateral_duty",
            "Collateral",
        )
    )


def pick_summer_duty(row: dict[str, str]) -> str:
    raw = pick_field(
        row,
        "Summer Training '26",
        "Summer Training",
        "Summer Duty",
        "summer_training",
    )
    return format_summer_duty(raw)


def format_summer_duty(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    text = re.sub(r"P\.S\.", "Plebe Summer", text, flags=re.I)
    text = re.sub(r"\bLNK\b", "Leatherneck", text, flags=re.I)
    return format_display_text(text)


def is_departed_row(row: dict[str, str], billet: str, name: str) -> bool:
    return bool(DEPARTED_RE.search(billet) or DEPARTED_RE.search(name))


def format_display_name(name: str) -> str:
    if "," not in name:
        return format_display_text(name)
    last, first = name.split(",", 1)
    return format_display_text(f"{first.strip()} {last.strip()}")


def compact_name_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def parse_roster_name(name: str) -> dict[str, str]:
    name = name.strip()
    if "," in name:
        last, first = name.split(",", 1)
        first = first.strip().rstrip(".")
        first_words = re.findall(r"[A-Za-z]+", first)
    else:
        last, first_words = name, re.findall(r"[A-Za-z]+", name)
    last_token = re.split(r"\s+", last.strip())[0]
    return {
        "last_key": compact_name_key(last_token),
        "first_initial": first_words[0][0].upper() if first_words else "",
        "first_word": first_words[0].upper() if first_words else "",
        "name": name,
    }


def parse_mentor_reference(ref: str) -> dict[str, str]:
    ref = ref.strip()
    if "," in ref:
        last, first = ref.split(",", 1)
        first = first.strip().rstrip(".")
        first_words = re.findall(r"[A-Za-z]+", first)
        if len(first) == 1:
            return {
                "last_key": compact_name_key(last),
                "first_initial": first.upper(),
                "first_word": "",
            }
        return {
            "last_key": compact_name_key(last),
            "first_initial": first_words[0][0].upper() if first_words else "",
            "first_word": first_words[0].upper() if first_words else "",
        }
    return {"last_key": compact_name_key(ref), "first_initial": "", "first_word": ""}


def last_name_matches(mentor_last: str, roster_last: str) -> bool:
    if not mentor_last or not roster_last:
        return False
    if mentor_last == roster_last:
        return True
    return mentor_last in roster_last or roster_last.startswith(mentor_last)


def match_mentor_reference(ref: str, marines: list[dict]) -> dict | None:
    if not ref.strip():
        return None
    criteria = parse_mentor_reference(ref)
    candidates: list[dict] = []
    for marine in marines:
        parsed = parse_roster_name(marine["name"])
        if not last_name_matches(criteria["last_key"], parsed["last_key"]):
            continue
        if criteria["first_initial"] and parsed["first_initial"] != criteria["first_initial"]:
            continue
        if criteria["first_word"] and parsed["first_word"] != criteria["first_word"]:
            continue
        candidates.append(marine)
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1 and criteria["first_word"]:
        narrowed = [
            m
            for m in candidates
            if parse_roster_name(m["name"])["first_word"] == criteria["first_word"]
        ]
        if len(narrowed) == 1:
            return narrowed[0]
    return candidates[0] if len(candidates) == 1 else None


def load_company_mentor_list() -> list[dict[str, str]]:
    path = DATA_DIR / "company-mentor-list.csv"
    if not path.exists():
        return []
    rows = read_csv(path)
    cleaned: list[dict[str, str]] = []
    for row in rows:
        company_raw = pick_field(row, "company", "Company")
        if not company_raw.isdigit():
            continue
        cleaned.append(
            {
                "company": company_raw,
                "ay26_mentor": pick_field(row, "ay26_mentor", "AY26 Marine Mentor", "ay26"),
                "ay27_mentor": pick_field(row, "ay27_mentor", "AY27 Marine Mentor", "ay27"),
            }
        )
    return cleaned


def build_mentors_from_company_list(marines: list[dict], config: dict[str, str]) -> tuple[list[dict], list[str]]:
    rows = load_company_mentor_list()
    if not rows:
        return [], []

    mentor_year = config.get("MENTOR_ACADEMIC_YEAR", "ay27").lower()
    report: list[str] = []
    mentors: list[dict] = []

    for row in rows:
        company = int(row["company"])
        ref = row.get(mentor_year) or row.get("ay27_mentor") or row.get("ay26_mentor")
        matched = match_mentor_reference(ref, marines)
        if matched:
            mentors.append(
                {
                    "company": company,
                    "battalion": (company - 1) // 6 + 1,
                    "name": format_display_name(matched["name"]),
                    "rank": format_rank(matched["rank"]),
                    "email": matched["email"],
                    "primaryDuty": matched.get("primaryDuty", ""),
                    "collateralDuty": matched.get("collateralDuty", ""),
                    "summerDuty": matched.get("summerDuty", ""),
                    "photo": f"{MENTOR_PHOTO_PREFIX}/company-{company:02d}.jpg",
                }
            )
        else:
            report.append(f"  Company {company}: no roster match for {ref!r}")
            mentors.append(
                {
                    "company": company,
                    "battalion": (company - 1) // 6 + 1,
                    "name": ref,
                    "rank": "",
                    "email": "",
                    "primaryDuty": "",
                    "collateralDuty": "",
                    "summerDuty": "",
                    "photo": f"{MENTOR_PHOTO_PREFIX}/company-{company:02d}.jpg",
                }
            )
    return mentors, report


def infer_community(
    *,
    department: str,
    billet: str,
    mentor_company: int,
    oe_rep: str,
) -> tuple[str, str]:
    dept = (department or "").strip().upper()
    billet_upper = (billet or "").strip().upper()

    if dept == "MARDET":
        return "mardet", COMMUNITY_LABELS["mardet"]
    if COMPANY_OFFICER_RE.search(billet_upper):
        return "company-officer", COMMUNITY_LABELS["company-officer"]
    if mentor_company:
        return "mentor", COMMUNITY_LABELS["mentor"]
    if oe_rep.strip() or dept == "PHYSICAL EDUCATION DEPT" or "PLAYER DEV" in billet_upper:
        return "coach", COMMUNITY_LABELS["coach"]
    if INSTRUCTOR_RE.search(billet_upper) or " DEPT" in dept:
        return "instructor", COMMUNITY_LABELS["instructor"]
    if dept == "COMMANDANT'S STAFF" or "COMMANDANT" in dept:
        return "support", COMMUNITY_LABELS["support"]
    return "other", COMMUNITY_LABELS["other"]


def build_public_notes(row: dict[str, str], community: str) -> str:
    explicit = pick_field(row, "notes", "Notes", "public_notes", "Public Notes")
    if explicit:
        return explicit
    parts: list[str] = []
    collateral = pick_field(
        row,
        "Collateral Duty '26",
        "Collateral Duty",
        "collateral_duty",
        "Collateral",
    )
    oe_rep = pick_field(row, "O/E Rep", "OE Rep", "Club", "Sport")
    if collateral:
        parts.append(f"Collateral duty: {collateral}")
    if oe_rep and community != "coach":
        parts.append(f"Club/sport: {oe_rep}")
    return " · ".join(parts)


def export_marines_csv_from_workbook(workbook_path: Path, csv_path: Path) -> bool:
    if not workbook_path.exists():
        return False
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl to export from workbook: pip install openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if MARINES_TAB not in wb.sheetnames:
        print(f"Workbook missing tab {MARINES_TAB!r}", file=sys.stderr)
        return False

    ws = wb[MARINES_TAB]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return False

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows[1:]:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            writer.writerow(["" if c is None else str(c).strip() for c in row[: len(headers)]])
    return True


def include_on_website(row: dict[str, str]) -> bool:
    flag = pick_field(
        row,
        "publish",
        "website",
        "show_on_website",
        "include_on_site",
        "public",
        "on_website",
    ).lower()
    if not flag:
        return True
    return flag in {"yes", "y", "true", "1", "x", "publish", "public"}


def slug_community(raw: str) -> str:
    key = raw.strip().lower()
    if key in COMMUNITY_SLUGS:
        return COMMUNITY_SLUGS[key]
    slug = re.sub(r"[^a-z0-9]+", "-", key).strip("-")
    return slug or "other"


def fetch_sheet_tab(config: dict[str, str], tab_name: str) -> list[dict[str, str]]:
    sheet_id = config.get("GOOGLE_SHEET_ID", "")
    creds_path = config.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not sheet_id or not creds_path or not tab_name:
        return []

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("Install Phase 2 deps: pip install -r scripts/requirements.txt", file=sys.stderr)
        sys.exit(1)

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(creds)
    worksheet = client.open_by_key(sheet_id).worksheet(tab_name)
    rows = worksheet.get_all_records()
    return [{str(k).strip(): "" if v is None else str(v).strip() for k, v in row.items()} for row in rows]


def get_rows(config: dict[str, str], csv_name: str, tab_config_key: str, default_tab: str) -> list[dict[str, str]]:
    csv_path = DATA_DIR / csv_name
    if not csv_path.exists() and csv_name == "marines.csv":
        workbook_path = DATA_DIR / WORKBOOK_FILENAME
        if export_marines_csv_from_workbook(workbook_path, csv_path):
            print(f"Exported {MARINES_TAB} → {csv_path}")
    rows = read_csv(csv_path)
    if rows:
        return rows
    tab = config.get(tab_config_key, default_tab)
    return fetch_sheet_tab(config, tab)


def normalize_marine(row: dict[str, str]) -> dict | None:
    if not include_on_website(row):
        return None
    name = pick_name(row)
    if not name:
        return None

    billet = pick_field(
        row,
        "billet",
        "Billet",
        "Billet / Primary Duty",
        "title",
        "Title",
        "position",
        "Position",
    )
    if is_departed_row(row, billet, name):
        return None

    department = pick_field(
        row,
        "department",
        "Department",
        "Cost / Staff Center",
        "org",
        "Org",
        "office",
        "Office",
    )
    mentor_company = parse_company(
        pick_field(row, "company", "Company", "Company Mentor", "co", "Co")
    )
    oe_rep = pick_field(row, "O/E Rep", "OE Rep", "Club", "Sport")
    company = mentor_company or company_from_billet(billet)

    community_raw = pick_field(
        row, "community", "Community", "type", "Type", "role", "Role", "category", "Category"
    )
    if community_raw:
        community = slug_community(community_raw)
        community_label = community_raw
    else:
        community, community_label = infer_community(
            department=department,
            billet=billet,
            mentor_company=mentor_company,
            oe_rep=oe_rep,
        )

    email = pick_field(row, "email", "Email").lower()
    primary_duty = format_primary_duty(billet, company)

    return {
        "name": name,
        "rank": format_rank(pick_field(row, "rank", "Rank", "grade", "Grade")),
        "community": community,
        "communityLabel": community_label,
        "billet": primary_duty or billet,
        "primaryDuty": primary_duty or billet,
        "collateralDuty": pick_collateral_duty(row),
        "summerDuty": pick_summer_duty(row),
        "location": pick_field(row, "location", "Location", "building", "Building", "where", "Where"),
        "department": department,
        "email": email,
        "phone": pick_field(row, "phone", "Phone", "office_phone", "Office Phone"),
        "company": company,
        "notes": build_public_notes(row, community),
    }


def normalize_mentor_from_marine(m: dict) -> dict | None:
    if m["community"] != "mentor" or not m["company"]:
        return None
    return {
        "company": m["company"],
        "battalion": (m["company"] - 1) // 6 + 1,
        "name": format_display_name(m["name"]),
        "rank": format_rank(m["rank"]),
        "email": m["email"],
        "primaryDuty": m.get("primaryDuty", m.get("billet", "")),
        "collateralDuty": m.get("collateralDuty", ""),
        "summerDuty": m.get("summerDuty", ""),
        "photo": f"{MENTOR_PHOTO_PREFIX}/company-{m['company']:02d}.jpg",
    }


def normalize_mentor_row(row: dict[str, str]) -> dict | None:
    if not include_on_website(row):
        return None
    company = int(pick_field(row, "company", "Company") or "0")
    if not company:
        return None
    photo = pick_field(row, "photo", "Photo") or f"{MENTOR_PHOTO_PREFIX}/company-{company:02d}.jpg"
    if photo and not photo.startswith("../") and not photo.startswith("http"):
        photo = f"{MENTOR_PHOTO_PREFIX}/{photo}"
    return {
        "company": company,
        "battalion": int(pick_field(row, "battalion", "Battalion") or str((company - 1) // 6 + 1)),
        "name": format_display_name(pick_name(row)),
        "rank": format_rank(pick_field(row, "rank", "Rank")),
        "email": pick_field(row, "email", "Email"),
        "primaryDuty": format_primary_duty(
            pick_field(row, "primaryDuty", "Primary Duty", "billet", "Billet"),
            company,
        ),
        "collateralDuty": pick_field(row, "collateralDuty", "Collateral Duty", "collateral"),
        "summerDuty": format_summer_duty(pick_field(row, "summerDuty", "Summer Duty", "summer_training")),
        "photo": photo,
    }


def js_string(value: str) -> str:
    return json.dumps(value or "")


def generate_marines_js(marines: list[dict]) -> str:
    lines = [
        "/** Auto-generated by scripts/sync-from-sheets.py — from Marines sheet / data/marines.csv */",
        "window.MARINES_ON_THE_YARD = [",
    ]
    for m in marines:
        lines.append(
            "  { "
            f"name: {js_string(m['name'])}, rank: {js_string(m['rank'])}, "
            f"community: {js_string(m['community'])}, communityLabel: {js_string(m['communityLabel'])}, "
            f"billet: {js_string(m['billet'])}, location: {js_string(m['location'])}, "
            f"department: {js_string(m['department'])}, email: {js_string(m['email'])}, "
            f"phone: {js_string(m['phone'])}, company: {m['company']}, "
            f"notes: {js_string(m['notes'])} }},"
        )
    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def generate_mentors_js(mentors: list[dict]) -> str:
    lines = [
        "/** Auto-generated by scripts/sync-from-sheets.py — do not edit by hand */",
        "window.COMPANY_MENTORS = [",
    ]
    for m in sorted(mentors, key=lambda x: x["company"]):
        lines.append(
            f"  {{ company: {m['company']}, battalion: {m['battalion']}, "
            f"name: {js_string(m['name'])}, rank: {js_string(m['rank'])}, "
            f"email: {js_string(m['email'])}, "
            f"primaryDuty: {js_string(m.get('primaryDuty', ''))}, "
            f"collateralDuty: {js_string(m.get('collateralDuty', ''))}, "
            f"summerDuty: {js_string(m.get('summerDuty', ''))}, "
            f"photo: {js_string(m['photo'])} }},"
        )
    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def content_hash(payload: dict) -> str:
    return hashlib.sha256(json.dumps(payload, sort_keys=True, ensure_ascii=True).encode()).hexdigest()


def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {}


def save_state(state: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2) + "\n")


def send_notification(config: dict[str, str], message: str) -> None:
    if config.get("SKIP_NOTIFY") == "1":
        return
    url = config.get("NOTIFY_WEBHOOK_URL", "")
    if not url:
        print("Notification skipped (set NOTIFY_WEBHOOK_URL in scripts/config.env)")
        return
    body = json.dumps({"text": message, "message": message, "title": "MARDET data updated"}).encode()
    req = urllib.request.Request(url, data=body, method="POST", headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            print(f"Notification sent ({resp.status})")
    except urllib.error.URLError:
        req = urllib.request.Request(url, data=message.encode(), method="POST")
        with urllib.request.urlopen(req, timeout=15) as resp:
            print(f"Notification sent ({resp.status})")


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Google Sheets / CSV data into website files")
    parser.add_argument("--dry-run", action="store_true", help="Show changes without writing files")
    args = parser.parse_args()

    config = load_config()
    report_lines = [f"MARDET sync — {datetime.now(timezone.utc).isoformat()}"]
    any_data = False
    any_changed = False
    state = load_state()
    combined_hash_input: dict = {}

    # --- Marines sheet (primary for Marines on the Yard) ---
    marines_raw = get_rows(config, "marines.csv", "SHEET_TAB_MARINES", "Marines")
    marines = [m for r in marines_raw if (m := normalize_marine(r))]
    if marines:
        any_data = True
        combined_hash_input["marines"] = marines
        marines_hash = content_hash({"marines": marines})
        marines_changed = marines_hash != state.get("marines_hash", "")
        any_changed = any_changed or (marines_changed and bool(state.get("marines_hash")))
        report_lines.append(f"Marines (Yard directory): {len(marines)} rows published")
        report_lines.append(f"  Marines changed: {marines_changed}")
        if not args.dry_run:
            INTRANET_JS_DIR.mkdir(parents=True, exist_ok=True)
            (INTRANET_JS_DIR / "marines-on-the-yard-data.js").write_text(generate_marines_js(marines))
            report_lines.append("  Wrote js/intranet/marines-on-the-yard-data.js")
        state["marines_hash"] = marines_hash
        state["marine_count"] = len(marines)
    else:
        report_lines.append("Marines: no data (export data/marines.csv from your Marines tab)")

    # --- Company mentors: assignment list + roster match, or legacy CSV / derived rows ---
    mentors: list[dict] = []
    mentor_report: list[str] = []
    list_mentors, list_report = build_mentors_from_company_list(marines, config)
    if list_mentors:
        mentors = list_mentors
        mentor_report = list_report
        report_lines.append(
            f"Company mentors: {len(mentors)} companies from data/company-mentor-list.csv "
            f"({config.get('MENTOR_ACADEMIC_YEAR', 'ay27').upper()})"
        )
        if mentor_report:
            report_lines.extend(mentor_report)
    else:
        mentors_raw = get_rows(config, "company-mentors.csv", "SHEET_TAB_COMPANY_MENTORS", "Company Mentors")
        if mentors_raw:
            mentors = [m for r in mentors_raw if (m := normalize_mentor_row(r))]
        elif marines:
            mentors = [m for row in marines if (m := normalize_mentor_from_marine(row))]
        if mentors:
            report_lines.append(f"Company mentors: {len(mentors)} rows (legacy source)")

    if mentors:
        any_data = True
        combined_hash_input["mentors"] = mentors
        mentors_hash = content_hash({"mentors": mentors})
        mentors_changed = mentors_hash != state.get("mentors_hash", "")
        any_changed = any_changed or (mentors_changed and bool(state.get("mentors_hash")))
        report_lines.append(f"  Mentors changed: {mentors_changed}")
        if not args.dry_run:
            INTRANET_JS_DIR.mkdir(parents=True, exist_ok=True)
            (INTRANET_JS_DIR / "company-mentors-data.js").write_text(generate_mentors_js(mentors))
            report_lines.append("  Wrote js/intranet/company-mentors-data.js")
        state["mentors_hash"] = mentors_hash
        state["mentor_count"] = len(mentors)

    if not any_data:
        print("No data found. Export data/marines.csv from your Marines tab (see data/README.md).")
        return 1

    if any_changed:
        report_lines.extend([
            "",
            "Review intranet pages after sync:",
            "  - pages/intranet/marines-on-the-yard.html",
            "  - pages/intranet/company-mentors.html",
        ])

    if args.dry_run:
        print("\n".join(report_lines))
        print("\nDry run — no files written.")
        return 0

    REPORT_FILE.write_text("\n".join(report_lines) + "\n")
    print("\n".join(report_lines))
    state["last_sync"] = datetime.now(timezone.utc).isoformat()
    save_state(state)

    if any_changed:
        send_notification(
            config,
            "MARDET sheet data changed. Run Cursor on usna-marine-det and review data/sync-report.txt",
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
